from openpyxl import Workbook
import argparse
import time
import pandas as pd
from pyecharts.charts import Line, Grid
from pyecharts import options as opts


def summary_num(params):
    """
    计算每个事件类型的触发总数、百公里触发次数、高速场景下触发数量、高速场景下带有回传信息的触发数量、
    城市场景下触发数量、城市场景下带有回传信息的触发数量，并将结果写入文件。

    Args:
        f (TextIOWrapper): 文件对象，用于写入结果

    Returns:
        None
    """
    sheet = params["sheet"][1]
    dt = {}
    data = params["data"]
    sheet.append(["汇总"])
    for i in set(data["事件类型"]):
        da = data[data["事件类型"] == i]
        sm = len(da)
        city = da[da["驾驶域"] == "urban"]

        city_yes = len(city[city["实时回传链接"] == "issueFinder"])
        speed = da[da["驾驶域"] == "highway"]
        speed_yes = len(speed[speed["实时回传链接"] == "issueFinder"])
        km3 = round(sm / params["km2"] * 100, 3)

        dt.update({
            i: [sm, km3, len(speed), speed_yes, len(city), city_yes]
        })

    sheet.append(["事件", "触发器", "总数", "百公里触发次数", "高速全部", "高速有回传", "城市全部", "城市有回传"])
    x = []
    for i in params["exc"]:
        exq = i[0]
        if exq in dt.keys():
            par = [exq, i[1]]
            par.extend(dt[exq])
            sheet.append(par)
        else:
            sheet.append([i[0], i[1], 0, 0, 0, 0, 0, 0])
    sheet.append(["", "自动驾驶里程", params["km2"]])
    sheet.append(["\n\n"])
    version_master(params)


def mb_km(km):
    """
    计算自动驾驶里程的总和。

    Args:
        km (pd.DataFrame): 包含自动驾驶里程信息的DataFrame，其中需要包含"version"和"自动驾驶里程"两列。

    Returns:
        Tuple[float, float]: 一个元组，包含两个浮点数，分别表示master版本和rb版本的自动驾驶里程总和。

    """
    master = km[km["version"].str.find("8.3.4") >= 0]
    master_km = master["自动驾驶里程（km）"]
    master_sum = master_km.astype(float).sum()
    rb_km = km[(km["version"].str.find("8.4.40") >= 0) | (km["version"].str.find("8.4.37") >= 0) |
               (km["version"].str.find("8.4.39") >= 0)]
    rb_km = rb_km["自动驾驶里程（km）"]
    rb_sum = rb_km.astype(float).sum()
    print("master:", int(master_sum))
    print("rb:", int(rb_sum))
    return master_sum, rb_sum


def version_num(params):
    """
    根据事件类型和版本信息，统计每个事件类型在各个版本中的数量，并写入到指定文件中。

    Args:
        f: 文件对象，用于写入统计结果。
        params: 字典类型，包含以下字段：
            - data: pandas DataFrame 类型，包含版本和事件类型等信息的数据集。
            - exc: 列表类型，包含需要统计的事件类型名称。

    Returns:
        vr: 字典类型，包含版本列表和每个事件类型对应的版本数量信息。
            字典的键为事件类型名称，值为一个二维列表，每个子列表包含两个元素：版本号和该事件类型在该版本中的数量。

    """
    # f = open("2.csv", "w")
    sheet = params["sheet"][0]
    vr = {}
    data = params["data"]
    title = params["exc"]
    versi = set(data["版本"])
    version_list = sorted(versi, key=lambda x: (int(x.split(".")[1]), int(x.split(".")[2]), int(x.split(".")[3])))
    vr["ver_list"] = version_list
    for i in set(data["事件类型"]):
        ex = data[data["事件类型"] == i]
        a = ex["版本"].to_list()
        for ver in version_list:
            num = ex[ex["版本"] == ver]
            if ver in a:
                try:
                    vr[i].append([ver, len(num)])
                except Exception as e:
                    vr[i] = []
                    vr[i].append([ver, len(num)])
            else:
                try:
                    vr[i].append([ver, 0])
                except Exception as e:
                    vr[i] = []
                    vr[i].append([ver, 0])
    c = len(vr["CIPV_VEH_LOST"])
    a = 0
    zero = ["0" for m in range(0, c)]
    for i in title:
        ex = i[0]
        if ex in vr.keys():

            version = list(map(lambda x: x[0], vr[ex]))
            num_ver = list(map(lambda x: x[1], vr[ex]))
            if a == 0 and version:
                par = ["版本\事件"]
                par.extend(version)
                sheet.append(par)
                a += 1
            par = [ex]
            par.extend(num_ver)
            sheet.append(par)
        else:
            par = [ex]
            par.extend(zero)
            sheet.append(par)
    return vr


def version_master(params):
    """
    计算master & rb 版本各事件类型的触发频率和总数，并写入文件。

    Args:
        f (TextIO): 文件对象，用于写入结果。

    Returns:
        None
    """
    sheet = params["sheet"][1]
    kk = params["kk"]
    for m, t in enumerate([params["master"], params["rb"]]):
        dt = {}
        if m == 0:
            sheet.append(["Master"])
        else:
            sheet.append(["RB"])
        sheet.append(["事件", "触发器", "总数", "百公里触发次数", "高速全部", "高速有回传", "城市全部", "城市有回传"])
        for i in set(t["事件类型"]):
            da = t[t["事件类型"] == i]
            sm = len(da)
            city = da[da["驾驶域"] == "urban"]

            city_yes = len(city[city["实时回传链接"] == "issueFinder"])
            speed = da[da["驾驶域"] == "highway"]
            speed_yes = len(speed[speed["实时回传链接"] == "issueFinder"])
            km3 = round(sm / kk[m] * 100, 3)

            dt.update({
                i: [sm, km3, len(speed), speed_yes, len(city), city_yes]
            })

        for i in params["exc"]:
            exq = i[0]
            exc = i[1]
            if exq in dt.keys():
                par = [exq, exc]
                par.extend(dt[exq])
                sheet.append(par)
                # f.write("%s,%s,%s\n" % (exq, exc, str(dt[exq]).strip("[]")))
            else:
                sheet.append([exq, exc, 0, 0, 0, 0, 0, 0])
                # f.write(f"{exq},{exc},0,0,0,0,0,0\n")
        sheet.append(["", "自动驾驶里程（km）", int(kk[m])])
        sheet.append(["\n\n"])


def rende(data: dict, save_name):
    """
    根据传入的数据绘制版本触发次数的折线图，并将图表渲染为HTML文件。

    Args:
        data (dict): 包含版本号和对应触发次数的字典，其中"ver_list"键对应的值为版本号的列表，其余键对应的值为版本号对应触发次数的列表（每个列表中的元素为元组，第一个元素为版本号，第二个元素为触发次数）。

    Returns:
        None: 该函数无返回值，将绘制的图表渲染为HTML文件。

    """
    line = Line()
    # 添加X轴和Y轴的数据
    ver_list = list(filter(lambda x: x.find("8.4.40") >= 0, data["ver_list"]))
    line.add_xaxis(ver_list)
    for i, j in data.items():
        if i == "ver_list":
            continue
        num = list(filter(lambda x: x[0].find("8.4.40") >= 0, j))
        num = list(map(lambda x: x[1], num))
        line.add_yaxis(f"{i}", num, linestyle_opts=opts.LineStyleOpts(width=3),
                       label_opts=opts.LabelOpts(is_show=False))

    line.set_global_opts(title_opts=opts.TitleOpts(title="版本触发次数", pos_left="center"),
                         legend_opts=opts.LegendOpts(orient='vertical', pos_right='center', pos_top='75%'),
                         xaxis_opts=opts.AxisOpts(
                             splitline_opts=opts.SplitLineOpts(is_show=False),
                             axislabel_opts=opts.LabelOpts(interval=0, rotate=-90)
                         ))
    g = Grid()
    g.add(line, grid_opts=opts.GridOpts(pos_bottom="40%"))
    # 渲染图表为HTML文件
    g.render(save_name)


def write_data(func):
    pass


def public_params(file):
    """
    获取公共参数，包括版本号、里程数和驾驶域。

    Returns:
        dict: 包含版本号、里程数和驾驶域的字典。
    """
    exc = [["CIPV_VEH_LOST", "车辆CIPV丢失触发器"],
           ["CIPV_POS_EXCEPTION_VEH", "车辆CIPV位置异常触发器"],
           ["CIPV_POS_SHIFT_DIFF_VEH", "车辆CIPV位置速度自洽性异常触发器"],
           ["CIPV_THETA_DIFF_VEH", "CIPV车辆朝向异常"],
           ["LARGE_VEHICLE_POS_EXCEPTION", "大车位置异常触发器"],
           ["LARGE_VEHICLE_POS_SHIFT_DIFF", "大车位置速度自洽性异常触发器"],
           ["LARGE_VEHICLE_THETA_DIFF", "大车朝向异常触发器"],
           ["NarrowCIPV", "narrow报出cipv"],
           ["RadarAssignFusionMining", "视觉ID跳变(短暂丢失)"],
           ["RadarTrackFusionMining", "视觉目标长时间丢失(CIPV"],
           ["UnderlyingVelocityFusionMining", "视觉速度灵敏性(CIPV)"],
           ["VRUVelocityFusionMining", "视觉速度朝向(行人)"],
           ["RadarPositionFusionMining", "视觉位置跳动(CIPV)"]]
    data = pd.read_excel(file, header=1)
    km = pd.read_excel(file, sheet_name="车辆里程信息", header=1)
    km2 = int(km[km["车号"] == " 合计"]["自动驾驶里程（km）"].str.replace(",", "").astype(float).iloc[0])
    wb = Workbook()
    wb.active.title = "summary"
    sum_sheet = wb.active
    version_num_sheet = wb.create_sheet("version_num")
    rb_sheet = wb.create_sheet("rb")
    master = data[data["版本"].str.find("8.3.4") >= 0]
    rb = data[(data["版本"].str.find("8.4.40") >= 0) | (data["版本"].str.find("8.4.37") >= 0) |
              (data["版本"].str.find("8.4.39") >= 0)]
    kk = mb_km(km)
    return {"exc": exc, "data": data, "km": km, "km2": km2, "master": master,
            "rb": rb, "kk": kk, "sheet": [version_num_sheet, sum_sheet, rb_sheet], "wb": wb}


def rb_km100(params):
    """
    根据给定的参数，统计特定版本的自动驾驶里程和特定事件类型在每百公里出现的次数，并将结果添加到给定的sheet中。

    Args:
        params (dict): 包含以下键值的字典：
            - sheet (list): 需要添加结果的二维列表。
            - rb (pandas.DataFrame): 包含路测数据的DataFrame，其中需包含“版本”列。
            - exc (list): 包含需要统计的事件类型列表，列表中每个元素为元组，元组的第一个元素为事件类型。
            - km (pandas.DataFrame): 包含自动驾驶里程的DataFrame，其中需包含“version”列和“自动驾驶里程（km）”列。
            - # car_id (可选，未使用): 车辆ID，本函数中未使用。

    Returns:
        None: 该函数将结果直接添加到给定的sheet中，不返回任何值。

    """

    sheet = params["sheet"][2]
    rb_data = params["rb"]
    exc = params["exc"]
    ver_list = [["1.4", "8.4.37"], ["1.5", "8.4.39"], ["2.0", "8.4.40"]]
    km = params["km"]
    # car_id = params["car_id"]
    for i in ver_list:
        rb_new = rb_data[rb_data["版本"].str.find(i[1]) >= 0]
        rb_km_data = km[km["version"].str.find(i[1]) >= 0]
        rb_km = rb_km_data["自动驾驶里程（km）"]
        rb_sum = rb_km.astype(float).sum()
        sheet.append([i[0], int(rb_sum)])
        print(f"{i[0]}:{int(rb_sum)}")
        for z in exc:
            ex = rb_new["事件类型"].unique()
            if z[0] in ex:
                rb_data_ex = rb_new[rb_new["事件类型"] == z[0]]
                ex_num = len(rb_data_ex)
                km100_num = round(ex_num / rb_sum * 100, 3)
                sheet.append([z[0], ex_num, km100_num])
            else:
                sheet.append([z[0], 0, 0])
        sheet.append(["\n\n"])


def args_parse():
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file", help="input file")
    parser.add_argument("-o", "--output", help="save file")
    parser.add_argument("-v", "--vis", help="line map")
    return parser.parse_args()


def save(save_file, params):
    summary_num(params)
    version_num(params)
    rb_km100(params)
    params["wb"].save(save_file)
    df = pd.read_excel(save_file, sheet_name="version_num")
    df = df.T
    with pd.ExcelWriter(save_file, engine='openpyxl', mode='a', if_sheet_exists="replace") as wri:
        df.to_excel(wri, sheet_name="version_num", header=False)


if __name__ == "__main__":
    with open("/Users/v_huangmin05/Downloads/1.txt", "r") as f:
        for i in f:
            file = "/Users/v_huangmin05/Downloads/" + i.strip()
            # file = f"/Users/v_huangmin05/Downloads/通用打点数据_1719217431301.xlsx"
            params = public_params(file)
            save_file = "1_" + file.split("/")[-1]
            save(save_file, params)
            print()
            time.sleep(1)
